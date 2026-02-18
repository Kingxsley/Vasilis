"""
Report Generation Service
Generates Excel and PDF reports for phishing campaigns and training
"""
import io
from datetime import datetime, timezone
from typing import List, Dict, Any
import logging

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib.enums import TA_CENTER, TA_LEFT

logger = logging.getLogger(__name__)


def generate_phishing_campaign_excel(campaign: dict, targets: list, stats: dict) -> bytes:
    """Generate Excel report for a phishing campaign"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Campaign Report"
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    stat_fill = PatternFill(start_color="D4A836", end_color="D4A836", fill_type="solid")
    clicked_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
    opened_fill = PatternFill(start_color="FFE066", end_color="FFE066", fill_type="solid")
    safe_fill = PatternFill(start_color="51CF66", end_color="51CF66", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Title
    ws.merge_cells('A1:G1')
    ws['A1'] = f"Phishing Campaign Report: {campaign.get('name', 'Unknown')}"
    ws['A1'].font = Font(bold=True, size=16)
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # Campaign Info
    ws['A3'] = "Campaign Details"
    ws['A3'].font = Font(bold=True, size=14)
    
    info_rows = [
        ("Organization ID:", campaign.get('organization_id', 'N/A')),
        ("Status:", campaign.get('status', 'N/A').upper()),
        ("Created:", campaign.get('created_at', 'N/A')),
        ("Started:", campaign.get('started_at', 'N/A') or 'Not Started'),
        ("Completed:", campaign.get('completed_at', 'N/A') or 'In Progress'),
    ]
    
    for i, (label, value) in enumerate(info_rows, start=4):
        ws[f'A{i}'] = label
        ws[f'A{i}'].font = Font(bold=True)
        ws[f'B{i}'] = str(value)
    
    # Statistics
    stats_row = 10
    ws[f'A{stats_row}'] = "Statistics"
    ws[f'A{stats_row}'].font = Font(bold=True, size=14)
    
    stat_headers = ['Total Targets', 'Emails Sent', 'Emails Opened', 'Links Clicked', 'Open Rate', 'Click Rate']
    stat_values = [
        stats.get('total_targets', 0),
        stats.get('emails_sent', 0),
        stats.get('emails_opened', 0),
        stats.get('links_clicked', 0),
        f"{stats.get('open_rate', 0)}%",
        f"{stats.get('click_rate', 0)}%"
    ]
    
    for i, header in enumerate(stat_headers):
        col = get_column_letter(i + 1)
        ws[f'{col}{stats_row + 1}'] = header
        ws[f'{col}{stats_row + 1}'].font = header_font
        ws[f'{col}{stats_row + 1}'].fill = stat_fill
        ws[f'{col}{stats_row + 1}'].alignment = Alignment(horizontal='center')
        ws[f'{col}{stats_row + 1}'].border = border
        
        ws[f'{col}{stats_row + 2}'] = stat_values[i]
        ws[f'{col}{stats_row + 2}'].alignment = Alignment(horizontal='center')
        ws[f'{col}{stats_row + 2}'].border = border
    
    # Targets Table
    targets_row = stats_row + 5
    ws[f'A{targets_row}'] = "Target Details"
    ws[f'A{targets_row}'].font = Font(bold=True, size=14)
    
    target_headers = ['Name', 'Email', 'Email Sent', 'Opened', 'Opened At', 'Clicked', 'Clicked At']
    header_row = targets_row + 1
    
    for i, header in enumerate(target_headers):
        col = get_column_letter(i + 1)
        ws[f'{col}{header_row}'] = header
        ws[f'{col}{header_row}'].font = header_font
        ws[f'{col}{header_row}'].fill = header_fill
        ws[f'{col}{header_row}'].alignment = Alignment(horizontal='center')
        ws[f'{col}{header_row}'].border = border
    
    for row_idx, target in enumerate(targets, start=header_row + 1):
        row_data = [
            target.get('user_name', 'N/A'),
            target.get('user_email', 'N/A'),
            'Yes' if target.get('email_sent') else 'No',
            'Yes' if target.get('email_opened') else 'No',
            target.get('email_opened_at', '') or '-',
            'Yes' if target.get('link_clicked') else 'No',
            target.get('link_clicked_at', '') or '-',
        ]
        
        for col_idx, value in enumerate(row_data):
            col = get_column_letter(col_idx + 1)
            ws[f'{col}{row_idx}'] = str(value)
            ws[f'{col}{row_idx}'].border = border
            ws[f'{col}{row_idx}'].alignment = Alignment(horizontal='center')
            
            # Color coding for status
            if col_idx == 5 and target.get('link_clicked'):  # Clicked column
                ws[f'{col}{row_idx}'].fill = clicked_fill
            elif col_idx == 3 and target.get('email_opened') and not target.get('link_clicked'):  # Opened
                ws[f'{col}{row_idx}'].fill = opened_fill
    
    # Adjust column widths
    column_widths = [20, 35, 12, 12, 20, 12, 20]
    for i, width in enumerate(column_widths):
        ws.column_dimensions[get_column_letter(i + 1)].width = width
    
    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def generate_phishing_campaign_pdf(campaign: dict, targets: list, stats: dict, org_name: str = None) -> bytes:
    """Generate PDF report for a phishing campaign"""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=0.5*inch, bottomMargin=0.5*inch)
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=colors.HexColor('#1F4E79'),
        alignment=TA_CENTER,
        spaceAfter=20
    )
    subtitle_style = ParagraphStyle(
        'Subtitle',
        parent=styles['Normal'],
        fontSize=12,
        textColor=colors.HexColor('#666666'),
        alignment=TA_CENTER,
        spaceAfter=30
    )
    section_style = ParagraphStyle(
        'Section',
        parent=styles['Heading2'],
        fontSize=14,
        textColor=colors.HexColor('#D4A836'),
        spaceBefore=20,
        spaceAfter=10
    )
    
    elements = []
    
    # Title
    elements.append(Paragraph("Phishing Campaign Report", title_style))
    elements.append(Paragraph(f"<b>{campaign.get('name', 'Unknown Campaign')}</b>", subtitle_style))
    if org_name:
        elements.append(Paragraph(f"Organization: {org_name}", subtitle_style))
    elements.append(Paragraph(f"Generated: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}", subtitle_style))
    
    elements.append(Spacer(1, 20))
    
    # Executive Summary
    elements.append(Paragraph("Executive Summary", section_style))
    
    summary_data = [
        ['Metric', 'Value', 'Assessment'],
        ['Total Targets', str(stats.get('total_targets', 0)), '-'],
        ['Emails Sent', str(stats.get('emails_sent', 0)), '-'],
        ['Emails Opened', str(stats.get('emails_opened', 0)), f"{stats.get('open_rate', 0)}% open rate"],
        ['Links Clicked', str(stats.get('links_clicked', 0)), f"{stats.get('click_rate', 0)}% click rate"],
    ]
    
    # Risk assessment
    click_rate = stats.get('click_rate', 0)
    if click_rate > 30:
        risk_level = "HIGH RISK - Immediate training recommended"
        risk_color = colors.HexColor('#FF6B6B')
    elif click_rate > 15:
        risk_level = "MEDIUM RISK - Additional training suggested"
        risk_color = colors.HexColor('#FFE066')
    else:
        risk_level = "LOW RISK - Good security awareness"
        risk_color = colors.HexColor('#51CF66')
    
    summary_data.append(['Risk Level', risk_level, ''])
    
    summary_table = Table(summary_data, colWidths=[120, 100, 200])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1F4E79')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 11),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#CCCCCC')),
        ('FONTSIZE', (0, 1), (-1, -1), 10),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F5F5F5')]),
    ]))
    elements.append(summary_table)
    
    elements.append(Spacer(1, 30))
    
    # Target Details
    elements.append(Paragraph("Individual Results", section_style))
    
    # Users who clicked (high risk)
    clicked_users = [t for t in targets if t.get('link_clicked')]
    opened_users = [t for t in targets if t.get('email_opened') and not t.get('link_clicked')]
    safe_users = [t for t in targets if not t.get('email_opened')]
    
    if clicked_users:
        elements.append(Paragraph("<font color='#FF6B6B'><b>Users Who Clicked (Require Training):</b></font>", styles['Normal']))
        click_data = [['Name', 'Email', 'Clicked At']]
        for u in clicked_users:
            click_data.append([
                u.get('user_name', 'N/A'),
                u.get('user_email', 'N/A'),
                str(u.get('link_clicked_at', 'N/A'))[:19]
            ])
        
        click_table = Table(click_data, colWidths=[120, 200, 150])
        click_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#FF6B6B')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#CCCCCC')),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
        ]))
        elements.append(click_table)
        elements.append(Spacer(1, 15))
    
    if opened_users:
        elements.append(Paragraph("<font color='#FFB300'><b>Users Who Opened (Caution):</b></font>", styles['Normal']))
        open_data = [['Name', 'Email', 'Opened At']]
        for u in opened_users:
            open_data.append([
                u.get('user_name', 'N/A'),
                u.get('user_email', 'N/A'),
                str(u.get('email_opened_at', 'N/A'))[:19]
            ])
        
        open_table = Table(open_data, colWidths=[120, 200, 150])
        open_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#FFB300')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#CCCCCC')),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
        ]))
        elements.append(open_table)
        elements.append(Spacer(1, 15))
    
    # Footer
    elements.append(Spacer(1, 30))
    footer_style = ParagraphStyle(
        'Footer',
        parent=styles['Normal'],
        fontSize=8,
        textColor=colors.HexColor('#999999'),
        alignment=TA_CENTER
    )
    elements.append(Paragraph("Generated by Vasilis NetShield Security Training Platform", footer_style))
    elements.append(Paragraph("This report is confidential and intended for authorized personnel only.", footer_style))
    
    doc.build(elements)
    buffer.seek(0)
    return buffer.getvalue()


def generate_training_report_excel(sessions: list, users_map: dict) -> bytes:
    """Generate Excel report for training sessions"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Training Report"
    
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Title
    ws.merge_cells('A1:G1')
    ws['A1'] = "Training Sessions Report"
    ws['A1'].font = Font(bold=True, size=16)
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # Headers
    headers = ['User Name', 'Email', 'Module', 'Status', 'Score', 'Started', 'Completed']
    for i, header in enumerate(headers):
        col = get_column_letter(i + 1)
        ws[f'{col}3'] = header
        ws[f'{col}3'].font = header_font
        ws[f'{col}3'].fill = header_fill
        ws[f'{col}3'].border = border
        ws[f'{col}3'].alignment = Alignment(horizontal='center')
    
    # Data
    for row_idx, session in enumerate(sessions, start=4):
        user = users_map.get(session.get('user_id'), {})
        row_data = [
            user.get('name', 'Unknown'),
            user.get('email', 'N/A'),
            session.get('module_id', '').replace('mod_', '').replace('_', ' ').title(),
            session.get('status', 'N/A').title(),
            f"{session.get('score', 0)}%",
            str(session.get('started_at', ''))[:19],
            str(session.get('completed_at', '-'))[:19] if session.get('completed_at') else '-',
        ]
        
        for col_idx, value in enumerate(row_data):
            col = get_column_letter(col_idx + 1)
            ws[f'{col}{row_idx}'] = value
            ws[f'{col}{row_idx}'].border = border
            ws[f'{col}{row_idx}'].alignment = Alignment(horizontal='center')
    
    # Adjust column widths
    column_widths = [20, 30, 25, 12, 10, 20, 20]
    for i, width in enumerate(column_widths):
        ws.column_dimensions[get_column_letter(i + 1)].width = width
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()
